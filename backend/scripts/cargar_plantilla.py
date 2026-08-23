"""Carga la plantilla de recolección en la base de datos y exporta el CSV.

La plantilla es un archivo de Excel que se llena a mano en el laboratorio. Este
script es el puente entre ese archivo y el modelo: lee la hoja de resumen,
comprueba que los datos tengan sentido, los guarda en la tabla `experiments` y
deja un CSV listo para entrenar.

Se ejecuta cada vez que llegan lotes nuevos:

    python backend/scripts/cargar_plantilla.py \
        --excel "Plantilla_recoleccion_GrillosIA.xlsx"

Sin `--base-datos` no toca PostgreSQL: solo valida y exporta el CSV, que es lo
que se necesita para entrenar. La base de datos es para conservar el histórico.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from ml.features import (  # noqa: E402
    DIETAS,
    ESPECIE_POR_DEFECTO,
    RANGOS,
    TARGET_COLUMNS,
)

HOJA_RESUMEN = "Registro de ensayos"
HOJA_DIETAS = "Codigos de dieta"

#: Traducción entre el código de variable de la plantilla y el nombre de la
#: columna en la tabla `experiments`. La plantilla puede reordenar columnas o
#: cambiar los títulos; mientras el código Vxx siga ahí, el cargador lo
#: encuentra.
POR_CODIGO = {
    "V01": "tipo_dieta",
    "V02": "alimento_g_dia",
    "V03": "temperatura",
    "V04": "humedad_ambiental",
    "V05": "fotoperiodo",
    "V06": "densidad",
    "V07": "especie",
    "V08": "tiempo_desarrollo",
    "V09": "n_grillos_inicio",
    "V10": "proteina_harina",
    "V11": "lipidos_harina",
    "V12": "tasa_supervivencia",
    "V13": "longitud_final",
    "V14": "biomasa_total",
}

#: Orden del CSV de salida. Las seis primeras son las entradas del modelo.
COLUMNAS_CSV = [
    "id_ensayo",
    "tipo_dieta",
    "alimento_g_dia",
    "temperatura",
    "humedad_ambiental",
    "especie",
    "tiempo_desarrollo",
    "proteina_harina",
    "lipidos_harina",
    "tasa_supervivencia",
    "biomasa_total",
    "longitud_final",
    "fotoperiodo",
    "densidad",
    "n_grillos_inicio",
    "fuente",
    "observaciones",
]

#: Sufijos que distinguen una columna auxiliar de la variable de verdad.
#: "V02a" es un insumo del cálculo; "V02" es el resultado.
AUXILIAR = re.compile(r"^V\d{2}[a-z]$")

#: Un título de variable lleva el código solo en su primera línea ("V03\n
#: Temperatura promedio"). La fila de grupos que va más arriba trae rangos
#: ("V03-V06 · Condiciones"), y esa no es la fila de títulos.
CODIGO = re.compile(r"^(V\d{2})[a-z]?$")


def _texto(celda) -> str:
    return "" if celda is None else str(celda).strip()


def localizar_encabezado(filas: list[tuple]) -> int:
    """Devuelve el índice de la fila de títulos.

    La plantilla lleva un par de filas de portada antes de la tabla, y esas
    filas cambian cuando alguien edita el archivo. En lugar de fijar el número
    de fila, buscamos la que contiene el código V01.
    """
    for i, fila in enumerate(filas):
        if any(
            CODIGO.match(_texto(c).split("\n")[0].strip()) for c in fila
        ):
            return i
    raise SystemExit(
        f"No se encontró la fila de títulos en la hoja '{HOJA_RESUMEN}'. "
        "Se busca la fila que contiene el código V01."
    )


def mapear_columnas(encabezado: tuple) -> dict[int, str]:
    """Asocia cada columna de la hoja con su nombre en la base de datos."""
    mapa: dict[int, str] = {}
    for i, celda in enumerate(encabezado):
        titulo = _texto(celda)
        if not titulo:
            continue
        if titulo.lower().startswith("id ensayo"):
            mapa[i] = "id_ensayo"
            continue
        if titulo.lower().startswith("observaciones"):
            mapa[i] = "observaciones"
            continue
        primera = titulo.split("\n")[0].strip()
        if AUXILIAR.match(primera):
            continue  # columna de apoyo, no es la variable
        m = CODIGO.match(primera)
        if m and m.group(1) in POR_CODIGO:
            mapa[i] = POR_CODIGO[m.group(1)]
    return mapa


def _numero(valor):
    if valor is None or _texto(valor) == "":
        return None
    try:
        return round(float(valor), 4)
    except (TypeError, ValueError):
        return None


NUMERICAS = {
    "alimento_g_dia",
    "temperatura",
    "humedad_ambiental",
    "fotoperiodo",
    "densidad",
    "n_grillos_inicio",
    "tiempo_desarrollo",
    "proteina_harina",
    "lipidos_harina",
    "tasa_supervivencia",
    "longitud_final",
    "biomasa_total",
}


def leer_lotes(ruta: Path) -> tuple[list[dict], list[str]]:
    """Lee la hoja de resumen. Devuelve las filas y los avisos encontrados."""
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "Falta openpyxl. Instálelo con:\n"
            "    pip install -r backend/requirements.txt"
        )

    libro = openpyxl.load_workbook(ruta, data_only=True)
    if HOJA_RESUMEN not in libro.sheetnames:
        raise SystemExit(
            f"El archivo no tiene la hoja '{HOJA_RESUMEN}'. "
            f"Hojas encontradas: {', '.join(libro.sheetnames)}"
        )

    hoja = libro[HOJA_RESUMEN]
    filas = list(hoja.iter_rows(values_only=True))
    inicio = localizar_encabezado(filas)
    mapa = mapear_columnas(filas[inicio])

    faltantes = set(POR_CODIGO.values()) - set(mapa.values())
    avisos: list[str] = []
    if faltantes:
        avisos.append(
            "Columnas de la plantilla que no se encontraron: "
            + ", ".join(sorted(faltantes))
        )

    lotes: list[dict] = []
    for fila in filas[inicio + 1 :]:
        registro = {col: fila[i] for i, col in mapa.items() if i < len(fila)}
        id_ensayo = _texto(registro.get("id_ensayo"))
        if not id_ensayo:
            continue  # fila en blanco al final de la hoja

        limpio: dict = {"id_ensayo": id_ensayo}
        for columna in COLUMNAS_CSV:
            if columna in ("id_ensayo", "fuente"):
                continue
            valor = registro.get(columna)
            if columna in NUMERICAS:
                limpio[columna] = _numero(valor)
            else:
                limpio[columna] = _texto(valor) or None

        # El tiempo de desarrollo y los grillos se cuentan en enteros.
        for entero in ("tiempo_desarrollo", "n_grillos_inicio"):
            if limpio.get(entero) is not None:
                limpio[entero] = int(limpio[entero])

        # Mientras la identificación taxonómica no esté cerrada, todos los
        # ejemplares se reportan a nivel de familia. Si la plantilla trae un
        # nombre de especie, se conserva en las observaciones y se avisa.
        especie = limpio.get("especie")
        if especie and especie != ESPECIE_POR_DEFECTO:
            avisos.append(
                f"{id_ensayo}: la plantilla dice especie '{especie}'; se "
                f"registra como '{ESPECIE_POR_DEFECTO}' hasta que la "
                "identificación esté confirmada."
            )
            nota = f"Especie según plantilla: {especie}."
            limpio["observaciones"] = (
                f"{limpio['observaciones']} {nota}".strip()
                if limpio.get("observaciones")
                else nota
            )
        limpio["especie"] = ESPECIE_POR_DEFECTO
        limpio["fuente"] = f"EXPERIMENTAL · {ruta.name}"

        lotes.append(limpio)

    return lotes, avisos


def revisar(lotes: list[dict]) -> list[str]:
    """Comprobaciones sobre los datos leídos.

    No detienen la carga: un lote sin análisis bromatológico es normal mientras
    los grillos siguen creciendo. Lo que hacen es dejar por escrito qué falta.
    """
    problemas: list[str] = []

    for lote in lotes:
        ident = lote["id_ensayo"]

        dieta = lote.get("tipo_dieta")
        if dieta and dieta not in DIETAS:
            problemas.append(
                f"{ident}: código de dieta '{dieta}' fuera del catálogo "
                f"({', '.join(DIETAS)})."
            )

        for columna, (minimo, maximo) in RANGOS.items():
            valor = lote.get(columna)
            if valor is not None and not (minimo <= valor <= maximo):
                problemas.append(
                    f"{ident}: {columna} = {valor} fuera del rango del "
                    f"estudio ({minimo}–{maximo})."
                )

        supervivencia = lote.get("tasa_supervivencia")
        if supervivencia is not None and not 0 <= supervivencia <= 100:
            problemas.append(
                f"{ident}: supervivencia = {supervivencia} %, imposible."
            )

    return problemas


def entrenables(lotes: list[dict]) -> list[dict]:
    """Lotes que ya sirven para entrenar: los que tienen el análisis hecho."""
    return [
        lote
        for lote in lotes
        if all(lote.get(t) is not None for t in TARGET_COLUMNS)
    ]


def exportar_csv(lotes: list[dict], destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("w", newline="", encoding="utf-8") as f:
        escritor = csv.DictWriter(f, fieldnames=COLUMNAS_CSV)
        escritor.writeheader()
        for lote in lotes:
            escritor.writerow({c: lote.get(c) for c in COLUMNAS_CSV})


def guardar_en_base(lotes: list[dict], url: str) -> int:
    """Vuelca los lotes en la tabla `experiments`.

    Reemplaza las filas cuya fuente sea la misma plantilla, para que volver a
    correr el script después de llenar más casillas actualice en lugar de
    duplicar.
    """
    from sqlalchemy import create_engine, delete
    from sqlalchemy.orm import Session

    from db.models import Base, Experiment

    motor = create_engine(url)
    Base.metadata.create_all(motor)

    # Columnas de la tabla que este cargador alimenta. `longitud_final` no
    # existe en `experiments`: la plantilla mide longitud corporal, no peso, y
    # no se debe hacer pasar una cosa por la otra.
    campos = [
        "tipo_dieta",
        "alimento_g_dia",
        "temperatura",
        "humedad_ambiental",
        "especie",
        "tiempo_desarrollo",
        "fotoperiodo",
        "densidad",
        "n_grillos_inicio",
        "proteina_harina",
        "lipidos_harina",
        "tasa_supervivencia",
        "biomasa_total",
        "observaciones",
        "fuente",
    ]

    with Session(motor) as sesion:
        fuentes = {lote["fuente"] for lote in lotes}
        for fuente in fuentes:
            sesion.execute(
                delete(Experiment).where(Experiment.fuente == fuente)
            )

        for lote in lotes:
            datos = {c: lote.get(c) for c in campos}
            nota = f"[{lote['id_ensayo']}]"
            datos["observaciones"] = (
                f"{nota} {datos['observaciones']}".strip()
                if datos.get("observaciones")
                else nota
            )
            sesion.add(Experiment(**datos))

        sesion.commit()

    return len(lotes)


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--excel", required=True, help="Plantilla de recolección")
    p.add_argument(
        "--csv",
        default="data/experimentos.csv",
        help="CSV consolidado de salida",
    )
    p.add_argument(
        "--base-datos",
        default=None,
        help=(
            "URL de PostgreSQL. Si se omite, no se toca la base de datos: "
            "solo se valida y se exporta el CSV."
        ),
    )
    args = p.parse_args()

    ruta = Path(args.excel)
    if not ruta.exists():
        raise SystemExit(f"No existe el archivo: {ruta}")

    lotes, avisos = leer_lotes(ruta)
    if not lotes:
        raise SystemExit("La plantilla no tiene ningún lote registrado.")

    problemas = revisar(lotes)
    listos = entrenables(lotes)

    print(f"\nPlantilla: {ruta.name}")
    print(f"Lotes registrados: {len(lotes)}")
    print(
        f"Lotes con análisis bromatológico: {len(listos)} "
        f"(los únicos que sirven para entrenar)"
    )

    if avisos:
        print("\nAvisos:")
        for a in avisos:
            print(f"  · {a}")

    if problemas:
        print("\nRevisar en la plantilla:")
        for x in problemas:
            print(f"  · {x}")

    destino = Path(args.csv)
    exportar_csv(lotes, destino)
    print(f"\nCSV consolidado: {destino}")

    if args.base_datos:
        n = guardar_en_base(lotes, args.base_datos)
        print(f"Guardados en la base de datos: {n} lotes")
    else:
        print("Base de datos: no se tocó (falta --base-datos)")

    if not listos:
        print(
            "\nTodavía no se puede entrenar con estos datos: ningún lote "
            "tiene proteína y lípidos de la harina. Mientras tanto se usa el "
            "conjunto simulado (backend/scripts/simular_datos.py)."
        )
    else:
        print(
            f"\nPara entrenar:\n"
            f"    python backend/scripts/entrenar.py --datos {destino}"
        )


if __name__ == "__main__":
    main()
